import openpyxl
import iter
from tkinter import *

def nomer(root,lbl,sheet):
    lbl.destroy()
    lbl = Label(root, text=f'Вкладка: {sheet}')
    lbl.place(x=20, y=150)
    root.update()

def start(filename, root, lbl, delta):
    file = filename
    wb1 = openpyxl.load_workbook(file,
                                 keep_vba=True)  # Открываем книгу. keep чтоб не было ошибки повреждения файла при сохранении
    cnt=0   #Счетчик пустых координат
    for sheet in range(14):  # для каждого листа по длине листов  len(wb1.sheetnames)
        nomer(root,lbl, sheet)
        if wb1.sheetnames[sheet] != 'Content' and wb1.sheetnames[sheet] != 'Protocol' and wb1.sheetnames[sheet] != 'GSM сканер' and wb1.sheetnames[sheet] != 'WCDMA сканер' and wb1.sheetnames[sheet] != 'LTE сканер':  # не содержащих Контент и Протокол
            wb1.active = sheet  # выбрать активный лист
            ws = wb1.active  # назвать его ws
            for row in range(3, ws.max_row):  # для строк с 3 по длину макс строк
                cnt=iter.colonka(row, 2, ws, cnt, delta)
                cnt=iter.colonka(row, 3, ws, cnt, delta)

    if cnt > 0:
        lbl.destroy()
        lbl = Label(root, text=f'Сохранение файла...')
        lbl.place(x=20, y=150)
        root.update()
        wb1.save(f'{file[0:-5]}-coordinats{file[-5:-1]}m')
        lbl.destroy()
        lbl = Label(root, text=f'Готово! Восстановлено координат: {cnt}')
        lbl.place(x=20, y=150)
    else:
        lbl.destroy()
        lbl = Label(root, text=f'Пустых координат не найдено. Создание файла не требуется!')
        lbl.place(x=20, y=150)
