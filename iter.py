from openpyxl.styles import Font, Alignment
def colonka(row, col:int, ws, cnt, delta):
    lat = ws.cell(row=row, column=col).value  # в lat значения второго столбца
    if lat == None:
        cnt+=1
        b = float(ws.cell(row=row - 1, column=col).value) + float(delta)
        ws.cell(row=row, column=col).value = b
    if ws.title == '0' or ws.title == 'SMS_text' or ws.title == '10':  # Если не 0 лист и не смс текст
        ws.cell(row=row, column=col).alignment = Alignment(vertical='top')  # выравнивает ячейку по верху
    if ws.title != '0':
        ws.cell(row=row, column=col).font = Font(name='Arial',
                                                   size='10')  # на всех, кроме листа 0 меняем шрифт
    return cnt

